import logging
import os
import datetime
from io import BytesIO
from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import ApplicationBuilder, CommandHandler, MessageHandler, filters, ContextTypes, ConversationHandler
from telegram.constants import ChatAction
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.chart import LineChart, Reference
from openpyxl.drawing.image import Image
from reportlab.platypus import SimpleDocTemplate, Paragraph, Table
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors
import matplotlib.pyplot as plt
from dotenv import load_dotenv

load_dotenv()

# Логирование
logger = logging.getLogger(__name__)
logging.basicConfig(format='%(asctime)s - %(name)s - %(levelname)s - %(message)s', level=logging.INFO)

# Константы из .env
BOT_TOKEN = os.getenv("BOT_TOKEN")
ADMIN_ID = int(os.getenv("ADMIN_ID", "8139881064"))
CHANNEL_ID = os.getenv("CHANNEL_ID", "@F_S_Ta")

CUR, NEW, COST, PERIOD = range(4)
HISTORY_FILE = "история.xlsx"

async def check_subscription(update: Update) -> bool:
    user_id = update.effective_user.id
    try:
        member = await update.get_bot().get_chat_member(CHANNEL_ID, user_id)
        return member.status in ["member", "administrator", "creator"]
    except:
        return False

async def start(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    if not await check_subscription(update):
        keyboard = [[InlineKeyboardButton("Подписаться", url=f"https://t.me/{CHANNEL_ID.lstrip('@')}")]]
        reply_markup = InlineKeyboardMarkup(keyboard)
        await update.message.reply_text(
            "Чтобы использовать бота, подпишитесь на канал:",
            reply_markup=reply_markup
        )
        return ConversationHandler.END

    await update.message.reply_text("Введите текущий тариф (₽/мес):")
    return CUR

async def cur_tariff(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    try:
        context.user_data['cur'] = float(update.message.text)
    except ValueError:
        await update.message.reply_text("Пожалуйста, введите число.")
        return CURэээээээ
    await update.message.reply_text("Введите а/плату для нового тарифа (₽/мес):")
    return NEW

async def new_tariff(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    try:
        context.user_data['new'] = float(update.message.text)
    except ValueError:
        await update.message.reply_text("Пожалуйста, введите число.")
        return NEW
    await update.message.reply_text("Введите стоимость подключения  (руб):")
    return COST

async def cost(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    try:
        context.user_data['cost'] = float(update.message.text)
    except ValueError:
        await update.message.reply_text("Пожалуйста, введите число.")
        return COST
    await update.message.reply_text("За какой период рассчитываем выгоду (лет)?")
    return PERIOD

async def generate_and_send_reports(cur, new, cost, period_years, user_name, bot, chat_id, admin_id=None):
    months = int(period_years * 12)
    cumulative_old = 0
    cumulative_new = cost
    payback_month = None
    rows = [["Месяц", "Старая", "Новая", "Экономия"]]

    for m in range(1, months + 1):
        cumulative_old += cur
        cumulative_new += new
        diff = cumulative_old - cumulative_new
        rows.append([m, round(cumulative_old), round(cumulative_new), round(diff)])
        if payback_month is None and diff >= 0:
            payback_month = m

    summary = f"Окупаемость: {payback_month} мес.\n" if payback_month else "Окупаемость не достигнута.\n"
    summary += f"Общая экономия за {months} мес.: {round(cumulative_old - cumulative_new)}₽"

    # PDF
    pdf_buffer = BytesIO()
    doc = SimpleDocTemplate(pdf_buffer, pagesize=A4)
    style = getSampleStyleSheet()["BodyText"]
    pdf_table = Table(rows)
    styles = [('BACKGROUND', (0, 0), (-1, 0), colors.lightblue),
              ('GRID', (0, 0), (-1, -1), 1, colors.grey),
              ('ALIGN', (0, 0), (-1, -1), 'CENTER')]
    for idx, row in enumerate(rows[1:], start=1):
        if row[3] < 0:
            styles.append(('TEXTCOLOR', (0, idx), (-1, idx), colors.red))
    pdf_table.setStyle(styles)
    intro_text = Paragraph(
        f"<b>Исходные данные:</b><br/>"
        f"Текущий тариф: {cur:.2f}₽/мес<br/>"
        f"Новый тариф: {new:.2f}₽/мес<br/>"
        f"Подключение: {cost:.2f}₽<br/>"
        f"Период: {months} мес.<br/><br/><b>Итоги:</b><br/>{summary.replace(chr(10), '<br/>')}",
        style)
    doc.build([intro_text, pdf_table])
    pdf_buffer.seek(0)
    await bot.send_document(chat_id=chat_id, document=pdf_buffer, filename="выгода.pdf")

    # Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Расчёт тарифа"
    bold = Font(bold=True)
    center = Alignment(horizontal="center")
    header_fill = PatternFill("solid", fgColor="BDD7EE")
    negative_fill = PatternFill("solid", fgColor="FFC7CE")
    ws.append(["Исходные данные"])
    ws.append(["Текущий тариф", cur])
    ws.append(["Новый тариф", new])
    ws.append(["Стоимость подключения", cost])
    ws.append(["Период (мес)", months])
    ws.append([])
    ws.append(rows[0])
    for col in range(1, 5):
        cell = ws.cell(row=7, column=col)
        cell.font = bold
        cell.fill = header_fill
        cell.alignment = center
    for idx, row in enumerate(rows[1:], start=1):
        ws.append(row)
        if row[3] < 0:
            for col in range(1, 5):
                ws.cell(row=7 + idx, column=col).fill = negative_fill
    chart = LineChart()
    chart.title = "График накопленных затрат"
    chart.y_axis.title = '₽'
    chart.x_axis.title = 'Месяц'
    data = Reference(ws, min_col=2, min_row=7, max_col=3, max_row=7 + months)
    cats = Reference(ws, min_col=1, min_row=8, max_row=7 + months)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    ws.add_chart(chart, "F8")
    plt.figure(figsize=(1, 1))
    plt.plot([0], [0], 'o', markersize=30, color='green')
    plt.axis('off')
    img_buf = BytesIO()
    plt.savefig(img_buf, format='png', bbox_inches='tight', transparent=True)
    img_buf.seek(0)
    logo_img = Image(img_buf)
    logo_img.width = 60
    logo_img.height = 60
    ws.add_image(logo_img, "F1")
    excel_buffer = BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    await bot.send_document(chat_id=chat_id, document=excel_buffer, filename="выгода.xlsx")

    # Сохраняем в историю
    now = datetime.datetime.now().strftime("%Y-%m-%d %H:%M")
    if not os.path.exists(HISTORY_FILE):
        wb_hist = Workbook()
        ws_hist = wb_hist.active
        ws_hist.title = "История"
        ws_hist.append(["Дата", "Пользователь", "Текущий тариф", "Новый тариф", "Подключение", "Период", "Окупаемость (мес)", "Экономия"])
    else:
        wb_hist = load_workbook(HISTORY_FILE)
        ws_hist = wb_hist.active
    ws_hist.append([
        now,
        user_name,
        cur, new, cost, months,
        payback_month if payback_month else "Не достигнута",
        round(cumulative_old - cumulative_new)
    ])
    wb_hist.save(HISTORY_FILE)

    # Отправка админу
    if admin_id:
        # PDF и Excel уже отправлены пользователю, поэтому сбрасываем курсоры
        pdf_buffer.seek(0)
        excel_buffer.seek(0)
        await bot.send_document(chat_id=admin_id, document=pdf_buffer, filename="user_выгода.pdf")
        await bot.send_document(chat_id=admin_id, document=excel_buffer, filename="user_выгода.xlsx")

async def period(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    try:
        context.user_data['period'] = float(update.message.text)
    except ValueError:
        await update.message.reply_text("Пожалуйста, введите число.")
        return PERIOD

    await generate_and_send_reports(
        cur=context.user_data['cur'],
        new=context.user_data['new'],
        cost=context.user_data['cost'],
        period_years=context.user_data['period'],
        user_name=update.effective_user.full_name,
        bot=context.bot,
        chat_id=update.effective_chat.id,
        admin_id=ADMIN_ID
    )
    return ConversationHandler.END

async def cancel(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    await update.message.reply_text('Расчёт отменён.')
    return ConversationHandler.END

async def history(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_name = update.effective_user.full_name
    if not os.path.exists(HISTORY_FILE):
        await update.message.reply_text("История пуста.")
        return
    wb_src = load_workbook(HISTORY_FILE)
    ws_src = wb_src.active
    wb_user = Workbook()
    ws_user = wb_user.active
    ws_user.title = "История"
    ws_user.append([cell.value for cell in ws_src[1]])
    for row in ws_src.iter_rows(min_row=2, values_only=True):
        if row[1] == user_name:
            ws_user.append(row)
    if ws_user.max_row == 1:
        await update.message.reply_text("У вас пока нет расчётов.")
        return
    buf = BytesIO()
    wb_user.save(buf)
    buf.seek(0)
    await update.message.reply_document(buf, filename="моя_история.xlsx")

async def help_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    help_text = (
        "Доступные команды:\n"
        "/start - Начать расчет тарифа\n"
        "/cancel - Отменить текущий ввод\n"
        "/history - Показать вашу историю расчетов\n"
        "/help - Показать это сообщение"
    )
    await update.message.reply_text(help_text)

if __name__ == '__main__':
    app = ApplicationBuilder().token(BOT_TOKEN).build()

    conv = ConversationHandler(
        entry_points=[CommandHandler('start', start)],
        states={
            CUR: [MessageHandler(filters.TEXT & ~filters.COMMAND, cur_tariff)],
            NEW: [MessageHandler(filters.TEXT & ~filters.COMMAND, new_tariff)],
            COST: [MessageHandler(filters.TEXT & ~filters.COMMAND, cost)],
            PERIOD: [MessageHandler(filters.TEXT & ~filters.COMMAND, period)],
        },
        fallbacks=[CommandHandler('cancel', cancel)]
    )

    app.add_handler(conv)
    app.add_handler(CommandHandler("history", history))
    app.add_handler(CommandHandler("help", help_command))
    app.run_polling()
        
